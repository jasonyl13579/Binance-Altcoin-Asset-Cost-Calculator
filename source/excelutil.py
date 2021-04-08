# -*- coding: utf-8 -*-
"""
Created on Tue Sep 15 18:16:30 2020

@author: Corn
"""

import sys
from os import sep, path
from pandas import DataFrame, ExcelWriter
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Border, Side, Font
import source.timeutil as tu

def color_negative_red(val):
    """
    Takes a scalar and returns a string with
    the css property `'color: red'` for negative
    strings, black otherwise.
    """
    color = 'red' if val < 0 else 'green'
    return 'color: %s' % color
def load_excel(filename, sheet_name, truncate_sheet):
    startrow = None
    writer = ExcelWriter(filename, engine='openpyxl')
    
    try:
        # try to open an existing workbook
        writer.book = load_workbook(filename)
        # get the last row in the existing Excel sheet
        # if it was not specified explicitly
        if startrow is None and sheet_name in writer.book.sheetnames:
            startrow = writer.book[sheet_name].max_row

        # truncate sheet
        if truncate_sheet and sheet_name in writer.book.sheetnames:
            # index of [sheet_name] sheet
            idx = writer.book.sheetnames.index(sheet_name)
            # remove [sheet_name]
            writer.book.remove(writer.book.worksheets[idx])
            # create an empty sheet [sheet_name] using old index
            writer.book.create_sheet(sheet_name, idx)

        # copy existing sheets
        writer.sheets = {ws.title:ws for ws in writer.book.worksheets}
    except FileNotFoundError:
        # file does not exist yet, we will create it
        pass
    return (startrow, writer)
def generate_single_report_to_excel(client, truncate_sheet=True):
    assets = client.assets
    accountname = client.accountname
    file_path = client.info.save_path
    if 'USDT' in assets: del assets['USDT']
    if 'USD' in assets: del assets['USD']
    iter_list = list(assets.keys())
    for asset in iter_list:
        if asset not in client.info.query_list:
            #print (asset)
            del assets[asset]
    client.info.asset_num = len(assets)
    
    source = [v.__dict__ for k, v in sorted(assets.items(), key=lambda d:d[0])]
    
    df = DataFrame(source)
    df['last_query_timestamp'].fillna(value=tu.DEFAULT_EPOCH_TIME, inplace=True)
    most_recent_date = df['last_query_timestamp'].max()
    df['last_query_timestamp'] = [tu.datetime_to_utc_time(x) for x in df['last_query_timestamp']]
    df['利潤'] = df['current_price'] * df['qty'] - df['quoteQty']
    df['USDT 價值'] = df['current_price'] * df['qty']
    
    df.rename(columns = {'name':'貨幣種類', 'cost': '價格成本', 'qty': '持有數量', 'quoteQty': '買進總成本', \
                         'current_price': '現價', 'transection_count': '交易次數', 'last_query_timestamp' : '最後交易時間'}, inplace = True)    
    df = df[['貨幣種類','利潤','價格成本','現價','買進總成本','USDT 價值','持有數量', '交易次數', '最後交易時間' ]]    
    df = df.sort_values('利潤', ascending=False)
    summary = DataFrame.from_dict( {"貨幣種類": ['TOTAL'],
                               "利潤":[df['利潤'].sum()],
                               "價格成本":[None],
                               "現價":[None],
                               "買進總成本":[None],
                               'USDT 價值':[None],
                               '持有數量':[None], 
                               '交易次數':[df['交易次數'].sum()], 
                               '最後交易時間':[tu.datetime_to_utc_time(most_recent_date)]}) 
    df = df.append(summary,ignore_index=True)
    
    
    filename = file_path + sep + '電子貨幣實時盈虧計算.xlsx'
    sheet_name = accountname
    (startrow, writer) = load_excel(filename, sheet_name, truncate_sheet)
    df.style.applymap(color_negative_red, subset = ['利潤']).to_excel(writer, sheet_name=sheet_name, float_format="%.4f")
    worksheet = writer.sheets[sheet_name] 
    for column_cells in worksheet.columns:
        worksheet.column_dimensions[column_cells[0].column_letter].width = 18

    writer.save() 
    writer.close()
    
    print ("Output report to " + filename)
'''
with ExcelWriter('output.xlsx') as writer:
    df.to_excel(writer, sheet_name='Sheet1', float_format="%.4f")
    writer.save()
'''

def generate_summary_report_to_excel(clients_dict, meta_file_path='./', truncate_sheet=False):
    
    
    for client in clients_dict.values():
        print (client)
    source = {0:'資產更新時間'}
    source_columns = ['資產更新時間']
    profit_data = []
    source_data = [tu.datetime_to_utc_time(tu.get_current_timestamp())]
    for client_name, client in sorted(clients_dict.items(),key=lambda d:d[1].info.index):
        profit_data.append(client_name + '_利潤')
        source_columns.append(client_name + '_利潤')
        source_columns.append(client_name + '_資產數量')
        source[client.info.index*2+1] = client_name + '_利潤'
        source[client.info.index*2+2] = client_name + '_資產數量'
        source_data.append(client.info.profit)
        source_data.append(client.info.asset_num)
    
    df = DataFrame([source_data], columns=source_columns)
    filename = meta_file_path + sep + '電子貨幣實時盈虧計算.xlsx'
    sheet_name = 'Summary'
    header = False
    (startrow, writer) = load_excel(filename, sheet_name, truncate_sheet)
    if startrow is None:
        startrow = 0
        header = True
    # write out the new sheet
    
    print (df, file=sys.stderr)
    df.style.applymap(color_negative_red, subset = profit_data).to_excel(writer, sheet_name=sheet_name, startrow=startrow, float_format="%.4f", index=False, header=header)
    ws= writer.sheets[sheet_name]
    thin = Side(border_style="thin", color="000000") 
    for index, column_cells in enumerate(ws.columns):
        cell = ws.cell(row=1, column=index+1)
        cell.value = source[index] if index in source else ''
        cell.font = Font(bold=True)
        cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)
        ws.column_dimensions[column_cells[0].column_letter].width = 24
    
    writer.save()
    writer.close()
    print ("Output summary to " + filename)
    